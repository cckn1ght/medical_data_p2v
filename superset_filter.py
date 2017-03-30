import json
# from apply_filter import sentence2phrases
import load_6546
from performance import performance
from pprs import string2words
from collections import Counter
from load_shared_task import load_shared_task
from gensim.models.phrases import Phraser, Phrases
import openpyxl
from remove_urls import remove_urls


def sentence2phrases(tweet_data, bigram, trigram):
    for t in tweet_data:
        t.text = trigram[bigram[t.text]]
    print('applied phrases model to our sentences')
    return tweet_data


def load_phrase_model():
    print('loading phrases model...')
    bigram = Phraser(Phrases.load('./cleaned_data/bigramPhraseModel'))
    trigram = Phraser(Phrases.load('./cleaned_data/trigramPhraseModel'))
    return bigram, trigram


def load_superset_similar_words(similar_words_file, topn=0, similarity=0.7, add_key=False):
    with open(similar_words_file, 'r') as f:
        similarity_dict = json.load(f)
    sim_words_set = set()

    for key, val in similarity_dict.items():
        if topn:
            words = {sim_word[0] for sim_word in val[:topn]}
            if add_key:
                words.add(key)
        if similarity:
            words = {sim_word[0]
                     for sim_word in val if float(sim_word[1]) >= similarity}
            if words and add_key:
                words.add(key)
        if (not topn) and (not similarity):
            words = {sim_word[0] for sim_word in val}
            if add_key:
                words.add(key)
        # words.add(key)
        sim_words_set |= words

    with open('./cleaned_data/similar_words.txt', 'w') as f:
        for word in sim_words_set:
            f.write(word + '\n')
    return sim_words_set


def super_sim_filter(phrase_tweet_data, topn=1, similarity=0.7, add_key=False):
    print('applying filter...')
    # sim_words_set = load_superset_similar_words(
    #     './cleaned_data/sim_transpose.json', topn=topn, similarity=similarity, add_key=add_key)
    sim_words_set = load_superset_similar_words(
        './cleaned_data/sim_transpose.json', topn=topn, similarity=similarity, add_key=add_key)
    print('There are {} similar words totally'.format(len(sim_words_set)))
    # filter sentences
    qualified_tweets = list()
    for tweet_d in phrase_tweet_data:
        # tweet_id = tweet_d.id
        sentence = tweet_d.text
        # tag = tweet_d.tag
        if not set(sentence).isdisjoint(sim_words_set):
            qualified_tweets.append(tweet_d)
            # qualified_sentences.append(sentence)
            # qualified_tags.append(tag)
            # qualified_ids.append(tweet_id)
    print('tweets left after filter: {}'.format(len(qualified_tweets)))
    print(Counter(t.tag for t in qualified_tweets))
    # print('No to Yes ratio after filter: {}'.format(
    #     Counter(qualified_tags)[0] / Counter(qualified_tags)[1]))
    return qualified_tweets


def apply_filter(bigram, trigram, sim_and_key, filename='6546', topn=0):
    wb = openpyxl.Workbook()
    sheet = wb.active
    col = 2
    for similarity, add_key in sim_and_key:
        if filename == '6546':
            tweet_data = load_6546.load(yes_only=False)
            tp_file = './test/Matrika/iteration2/tp_iteration2_superset_cleaned_data.csv'
            wb_file = './cleaned_data/6546_iteration2_result.xlsx'
            prediction_file = './cleaned_data/6546_prediction.csv'
            # fname = './test/Matrika/iteration2/6546_prediction_superset_cleaned_data_ST.csv'
        elif filename == '4819':
            tweet_data = load_shared_task(yes_only=False)
            tp_file = './test/4819_shared_task/tp_4819_superset_cleaned_data.csv'
            wb_file = './cleaned_data/4819_sharedtask_in_result.xlsx'
            prediction_file = './cleaned_data/4819_prediction.csv'
            # fname = './test/4819_shared_task/4819_prediction_superset_cleaned_data.csv'
        cleaned_tweet_data = remove_urls(tweet_data)
        for anno in cleaned_tweet_data:
            anno.text = string2words(anno.text, remove_stopwords=False)

        # sentences = [string2words(s, remove_stopwords=True)
        #              for s in raw_sentences]
        phrase_tweet_data = sentence2phrases(
            cleaned_tweet_data, bigram, trigram)
        filtered_tweet_data = super_sim_filter(
            phrase_tweet_data, topn=topn, similarity=similarity, add_key=add_key)
        print('--------------------------------------')
        Performance = performance(
            tweet_data, filtered_tweet_data, tp_file, prediction_file, similarity, add_key)
        sheet.cell(row=1, column=col).value = similarity
        sheet.cell(row=2, column=col).value = str(add_key)
        row = 3
        for val in Performance:
            sheet.cell(row=row, column=col).value = val
            row += 1
        col += 1
        # save_prediction(fname, id_truetag_predtag, raw_sentences)
    wb.save(wb_file)
    return Performance


def main():
    bigram, trigram = load_phrase_model()
    sim_and_key = [(0.8, False), (0.8, True), (0.75, False),
                   (0.75, True), (0.7, False), (0.7, True), (0.65, False), (0.65, True), (0.6, False), (0.6, True)]

    # sim_and_key = [(0.7, False)]
    apply_filter(bigram, trigram, sim_and_key, filename='4819',
                 topn=0)


if __name__ == '__main__':
    main()
